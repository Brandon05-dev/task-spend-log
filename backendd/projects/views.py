"""
Views for the expense tracker API.

This module contains Django REST Framework ViewSets for handling
API requests for projects and expenses, including statement generation.
"""

import io
from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

# PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Project, Expense
from .serializers import ProjectSerializer, ProjectSummarySerializer, ExpenseSerializer


class ProjectViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing projects.
    
    Provides CRUD operations for projects and includes custom actions
    for generating PDF and Excel statements.
    """
    queryset = Project.objects.all()
    
    def get_serializer_class(self):
        """
        Return appropriate serializer based on action.
        
        Use ProjectSummarySerializer for list view to avoid loading
        all expenses, and ProjectSerializer for detail view.
        """
        if self.action == 'list':
            return ProjectSummarySerializer
        return ProjectSerializer
    
    def list(self, request, *args, **kwargs):
        """
        List all projects with summary information.
        """
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'count': queryset.count(),
            'results': serializer.data
        })
    
    def retrieve(self, request, *args, **kwargs):
        """
        Retrieve a specific project with all its expenses.
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def statement(self, request, pk=None):
        """
        Generate PDF and Excel statements for a project.
        
        Query parameters:
        - format: 'pdf' or 'excel' (default: 'pdf')
        """
        project = get_object_or_404(Project, pk=pk)
        format_type = request.query_params.get('format', 'pdf').lower()
        
        if format_type == 'excel':
            return self._generate_excel_statement(project)
        else:
            return self._generate_pdf_statement(project)
    
    def _generate_pdf_statement(self, project):
        """
        Generate PDF statement for a project using ReportLab.
        """
        # Create a BytesIO buffer to hold the PDF
        buffer = io.BytesIO()
        
        # Create the PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Build the document content
        story = []
        
        # Title
        title = Paragraph("Expense Statement", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Project information
        project_info = [
            ['Project Name:', project.name],
            ['Description:', project.description or 'No description'],
            ['Created Date:', project.created_at.strftime('%B %d, %Y')],
            ['Total Expenses:', f'${project.total_expenses:.2f}'],
            ['Number of Expenses:', str(project.expense_count)],
        ]
        
        project_table = Table(project_info, colWidths=[2*inch, 4*inch])
        project_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        story.append(project_table)
        story.append(Spacer(1, 30))
        
        # Expenses table
        if project.expenses.exists():
            expenses_title = Paragraph("Expense Details", styles['Heading2'])
            story.append(expenses_title)
            story.append(Spacer(1, 12))
            
            # Table headers
            expense_data = [['Date', 'Description', 'Amount']]
            
            # Add expense rows
            for expense in project.expenses.all():
                expense_data.append([
                    expense.date.strftime('%Y-%m-%d'),
                    expense.description[:50] + ('...' if len(expense.description) > 50 else ''),
                    f'${expense.amount:.2f}'
                ])
            
            expense_table = Table(expense_data, colWidths=[1.5*inch, 3.5*inch, 1.5*inch])
            expense_table.setStyle(TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),  # Amount column right-aligned
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                
                # Data rows
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(expense_table)
        else:
            no_expenses = Paragraph("No expenses recorded for this project.", styles['Normal'])
            story.append(no_expenses)
        
        # Build PDF
        doc.build(story)
        
        # Get the value of the BytesIO buffer and create response
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{project.name}_statement.pdf"'
        response.write(pdf)
        
        return response
    
    def _generate_excel_statement(self, project):
        """
        Generate Excel statement for a project using openpyxl.
        """
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Statement"
        
        # Define styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=18)
        normal_font = Font(size=12)
        
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Title
        ws['A1'] = "Expense Statement"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment
        ws.merge_cells('A1:C1')
        
        # Project information
        ws['A3'] = "Project Name:"
        ws['B3'] = project.name
        ws['A4'] = "Description:"
        ws['B4'] = project.description or 'No description'
        ws['A5'] = "Created Date:"
        ws['B5'] = project.created_at.strftime('%B %d, %Y')
        ws['A6'] = "Total Expenses:"
        ws['B6'] = f'${project.total_expenses:.2f}'
        ws['A7'] = "Number of Expenses:"
        ws['B7'] = project.expense_count
        
        # Style project info
        for row in range(3, 8):
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'].font = normal_font
        
        # Expenses header
        if project.expenses.exists():
            ws['A9'] = "Expense Details"
            ws['A9'].font = title_font
            ws.merge_cells('A9:C9')
            
            # Table headers
            ws['A11'] = "Date"
            ws['B11'] = "Description"
            ws['C11'] = "Amount"
            
            for col in ['A11', 'B11', 'C11']:
                ws[col].font = header_font
                ws[col].fill = header_fill
            
            # Add expense data
            row = 12
            for expense in project.expenses.all():
                ws[f'A{row}'] = expense.date.strftime('%Y-%m-%d')
                ws[f'B{row}'] = expense.description
                ws[f'C{row}'] = float(expense.amount)
                
                # Style the row
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].font = normal_font
                
                row += 1
            
            # Format amount column as currency
            for r in range(12, row):
                ws[f'C{r}'].number_format = '$#,##0.00'
        else:
            ws['A9'] = "No expenses recorded for this project."
            ws['A9'].font = normal_font
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        
        # Save to BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{project.name}_statement.xlsx"'
        
        return response


class ExpenseViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing expenses.
    
    Provides CRUD operations for expenses.
    """
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer
    
    def list(self, request, *args, **kwargs):
        """
        List all expenses with optional project filtering.
        
        Query parameters:
        - project: Filter expenses by project ID
        """
        queryset = self.get_queryset()
        
        # Filter by project if specified
        project_id = request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'count': queryset.count(),
            'results': serializer.data
        })
